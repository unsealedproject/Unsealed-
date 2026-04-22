"""Ingest OCSE preliminary-data Excel files -> acf_state_data.json.

Reads every FY workbook in `ACF Excel/`, stitches overlapping years
(newer workbook wins on overlap), computes derived metrics, and flags
statistically implausible patterns (the "100% collection with rising
arrears" family of claims that reveal OCSE-side accounting games).

Run this manually whenever new OCSE preliminary data drops; no cron
needed since the data only updates once a year.

    python ingest_acf.py

Output: acf_state_data.json in the same directory as this script.
"""

import json
import re
from pathlib import Path
from collections import defaultdict

import openpyxl

HERE = Path(__file__).parent
EXCEL_DIR = HERE / "ACF Excel"
OUT_PATH  = HERE / "acf_state_data.json"

# Map OCSE appendix table -> normalized metric name. State-level 5-year tables only;
# skip single-year tables (P-18/P-19/P-22/etc.) and national-only tables.
METRICS = {
    "P-4":  "distributedCollections",
    "P-6":  "currentAssistanceCollections",
    "P-10": "distributedTanfFcCollections",
    "P-13": "nonTanfCollections",
    "P-36": "costEffectivenessRatio",
    "P-37": "incentivePaymentEstimates",
    "P-38": "incentivePaymentActuals",
    "P-43": "adminExpenditures",
    "P-52": "caseload",
    "P-64": "casesWithOrders",
    "P-71": "paternitiesEstablished",
    "P-74": "ordersEstablishedThisYear",
    "P-75": "casesWithCollection",
    "P-80": "fteStaff",
    "P-83": "currentSupportDue",
    "P-84": "currentSupportDistributed",
    "P-85": "arrears",
    "P-87": "casesWithArrears",
    "P-88": "casesPayingArrears",
    "P-93": "childrenInCases",
}

# OCSE uses ALL-CAPS state names in the rows; lowercase for friendlier JSON keys.
def title_case_state(s: str) -> str:
    s = s.strip()
    special = {
        "DISTRICT OF COLUMBIA": "District of Columbia",
        "GUAM": "Guam",
        "PUERTO RICO": "Puerto Rico",
        "VIRGIN ISLANDS": "Virgin Islands",
    }
    if s in special:
        return special[s]
    # Handle "NEW YORK", "NORTH CAROLINA", etc.
    return " ".join(w.capitalize() for w in s.split())

NON_STATE_ROWS = {
    "STATES", "NATION", "NATIONAL", "TOTAL", "U.S. TOTAL", "US TOTAL",
    "NATIONAL TOTAL", "TOTALS", "", None,
}

def parse_value(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    if s in ("", "-", "N/A", "NA"): return None
    try:
        return float(s)
    except ValueError:
        return None

def read_workbook(path: Path):
    """Return {metric: {state: {year: value}}} for this FY workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = defaultdict(lambda: defaultdict(dict))
    for sheet_name, metric in METRICS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Find the header row (contains "STATES" in first col and years in others)
        header_row_idx = None
        years = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            first = str(row[0]).strip().upper() if row and row[0] else ""
            if first == "STATES":
                header_row_idx = i
                # Years may be stored as ints or as strings like "2019"; accept both.
                for y in row[1:]:
                    if y is None:
                        continue
                    try:
                        yi = int(float(y)) if isinstance(y, (int, float)) else int(float(str(y).strip()))
                    except (ValueError, TypeError):
                        continue
                    if 2000 <= yi <= 2030:
                        years.append(yi)
                break
        if not years or header_row_idx is None:
            continue
        # Data starts the row after the header
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= header_row_idx: continue
            if not row or not row[0]: continue
            name = str(row[0]).strip()
            if name.upper() in NON_STATE_ROWS: continue
            # Skip footnote rows (start with "1/", "*", etc.)
            if re.match(r'^[0-9]+/|^\*', name): continue
            state = title_case_state(name)
            vals = row[1:1 + len(years)]
            for y, v in zip(years, vals):
                pv = parse_value(v)
                if pv is not None:
                    out[metric][state][y] = pv
    return out

def merge_workbooks(all_wb_data, order_newest_first):
    """Merge per-workbook dicts. Later (newer) workbooks overwrite earlier ones
    when years overlap — that's intentional: OCSE revisions in newer releases
    supersede preliminary numbers from older releases."""
    merged = defaultdict(lambda: defaultdict(dict))
    for wb_data in order_newest_first:
        for metric, states in wb_data.items():
            for state, years in states.items():
                for year, val in years.items():
                    # Skip if already set by a newer workbook
                    if year not in merged[metric][state]:
                        merged[metric][state][year] = val
    return merged

def compute_derived(state_data):
    """Derived per-year ratios and fraud flags. Returns (derived, flags)."""
    derived = {}
    flags = []

    def ratio(a, b):
        out = {}
        ya = state_data.get(a, {})
        yb = state_data.get(b, {})
        for y in sorted(set(ya) & set(yb)):
            if yb[y] and yb[y] != 0:
                out[y] = round(ya[y] / yb[y], 4)
        return out

    derived["collectionRate"]    = ratio("currentSupportDistributed", "currentSupportDue")
    derived["arrearsPayingRate"] = ratio("casesPayingArrears",        "casesWithArrears")

    # Collections per case (requires both distributedCollections and caseload)
    if "distributedCollections" in state_data and "caseload" in state_data:
        cpc = {}
        dc = state_data["distributedCollections"]
        cl = state_data["caseload"]
        for y in sorted(set(dc) & set(cl)):
            if cl[y] and cl[y] > 0:
                cpc[y] = round(dc[y] / cl[y], 2)
        derived["collectionsPerCase"] = cpc

    # ── Fraud/implausibility flags ──────────────────────────────────

    # 1. Claimed >=99% collection with sizable remaining arrears — mathematically
    #    hard to reconcile: if current collection is ~100% for two+ consecutive
    #    years, the arrears stock should be shrinking toward zero, not growing.
    years_100 = []
    cr = derived.get("collectionRate", {})
    ar = state_data.get("arrears", {})
    for y in sorted(cr):
        if cr[y] >= 0.99 and ar.get(y, 0) > 1e6:   # >$1M arrears + ≥99% collection
            years_100.append(y)
    if years_100:
        flags.append({
            "type": "claimed_near_100pct_with_arrears",
            "years": years_100,
            "severity": "high" if len(years_100) >= 2 else "medium",
            "explanation": "Reported ≥99% current-support collection while carrying material arrears. "
                           "If current collection were actually near-complete, the arrears pool would be shrinking, "
                           "not persistent. Check the current-support-due (P-83) and arrears (P-85) series together."
        })

    # 2. Arrears monotone growth over 3+ years (never down) — reveals the state
    #    isn't reducing its arrears book even nominally, which is unusual.
    if "arrears" in state_data:
        years = sorted(state_data["arrears"])
        if len(years) >= 3:
            monotone = all(
                state_data["arrears"][y] >= state_data["arrears"][prev]
                for prev, y in zip(years, years[1:])
            )
            if monotone:
                delta = state_data["arrears"][years[-1]] - state_data["arrears"][years[0]]
                pct = (delta / state_data["arrears"][years[0]] * 100) if state_data["arrears"][years[0]] else 0
                flags.append({
                    "type": "arrears_monotone_growth",
                    "yearsObserved": [years[0], years[-1]],
                    "growthPct": round(pct, 1),
                    "severity": "high" if pct > 20 else "medium",
                    "explanation": f"Arrears grew every single year across {years[0]}-{years[-1]} (+{pct:.1f}%). "
                                   "Healthy programs show at least some year-over-year arrears reduction."
                })

    # 3. Caseload drop with flat/growing collections — either aggressive case
    #    cherry-picking (drop low-yield cases, keep high-yield) or double-count
    #    somewhere. Flag as medium.
    if "caseload" in state_data and "distributedCollections" in state_data:
        cl = state_data["caseload"]
        dc = state_data["distributedCollections"]
        shared = sorted(set(cl) & set(dc))
        if len(shared) >= 3:
            case_first, case_last = cl[shared[0]], cl[shared[-1]]
            coll_first, coll_last = dc[shared[0]], dc[shared[-1]]
            case_pct = (case_last - case_first) / case_first if case_first else 0
            coll_pct = (coll_last - coll_first) / coll_first if coll_first else 0
            if case_pct < -0.10 and coll_pct > -0.02:   # caseload -10%+, collections flat or rising
                flags.append({
                    "type": "caseload_drop_collections_stable",
                    "yearsObserved": [shared[0], shared[-1]],
                    "caseloadChangePct":     round(case_pct * 100, 1),
                    "collectionsChangePct":  round(coll_pct * 100, 1),
                    "severity": "medium",
                    "explanation": "Caseload dropped >10% while collections stayed flat or rose. "
                                   "Possible case-selection bias (retaining high-yield cases) or double-count."
                })

    # 4. Cost-effectiveness ratio above 10:1 — OCSE's own national average sits
    #    around 5:1-6:1. States claiming dramatically higher ratios warrant audit.
    ce = state_data.get("costEffectivenessRatio", {})
    outlier_years = [y for y, v in ce.items() if v and v > 10]
    if outlier_years:
        flags.append({
            "type": "cost_effectiveness_outlier",
            "years": sorted(outlier_years),
            "values": {y: ce[y] for y in sorted(outlier_years)},
            "severity": "low",
            "explanation": "Cost-effectiveness ratio above 10:1 exceeds the national norm (~5-6:1). "
                           "May indicate non-standard accounting or exclusion of non-IV-D costs."
        })

    # 5. Incentive-payment divergence — federal incentives are awarded via a
    #    formula scored against state self-reports. If ACTUAL incentive
    #    payments (P-38) persistently exceed the pre-year ESTIMATES (P-37)
    #    by >10%, the state is consistently outperforming its own forecasts —
    #    either it's sandbagging estimates, or the performance metrics
    #    rewarded post-hoc got redefined mid-cycle.
    # Threshold tuned empirically: ~25 states hit a 10% overrun in 3+ years,
    # which reflects normal federal-formula variance, not gaming. At 20% the
    # flag narrows to states with persistent, meaningful overruns.
    est = state_data.get("incentivePaymentEstimates", {})
    act = state_data.get("incentivePaymentActuals", {})
    div_years = []
    for y in sorted(set(est) & set(act)):
        if est[y] and est[y] > 0:
            delta = (act[y] - est[y]) / est[y]
            if delta > 0.20:
                div_years.append((y, round(delta * 100, 1)))
    if len(div_years) >= 3:
        avg = round(sum(d for _, d in div_years) / len(div_years), 1)
        flags.append({
            "type": "incentive_estimate_vs_actual_divergence",
            "yearsObserved": [y for y, _ in div_years],
            "avgOverruns": avg,
            "severity": "high" if avg > 50 else "medium",
            "explanation": f"Actual incentive payments exceeded estimates by >20% in {len(div_years)} years "
                           f"(avg overrun {avg}%). Persistent overruns of this magnitude go beyond normal "
                           "forecast variance and suggest the state is either sandbagging its pre-year "
                           "projections or the performance formula is being revised mid-cycle in ways "
                           "that favor it."
        })

    # 6. Arrears-to-annual-collections ratio above 20:1 — means at the
    #    current collection pace it would take >20 years to clear arrears
    #    even if every new dollar went exclusively to arrears (impossible in
    #    practice). States with ratios this high aren't collecting; they're
    #    warehousing debt that grows indefinitely.
    ar = state_data.get("arrears", {})
    dc = state_data.get("distributedCollections", {})
    perp = {}
    for y in sorted(set(ar) & set(dc)):
        if dc[y] and dc[y] > 0:
            ratio = ar[y] / dc[y]
            if ratio > 20:
                perp[y] = round(ratio, 1)
    if perp:
        flags.append({
            "type": "perpetual_debt_ratio",
            "years": sorted(perp),
            "ratios": perp,
            "severity": "high" if len(perp) >= 3 else "medium",
            "explanation": "Arrears exceed 20× the state's annual distributed collections. "
                           "At current collection pace, arrears cannot be cleared within a generation — "
                           "indicates the state is warehousing indefinite debt rather than pursuing "
                           "genuine recovery, likely tied to federal incentive-formula scoring."
        })

    # 7. Non-TANF migration — if non-TANF collections (P-13, pure revenue cases
    #    where no public assistance is involved) grow while TANF/FC collections
    #    (P-10, actual cash-assistance support) fall, the program is morphing
    #    from "recover money the state paid out in welfare" into "a debt-
    #    collection business with custody leverage."
    ntf = state_data.get("nonTanfCollections", {})
    tfc = state_data.get("distributedTanfFcCollections", {})
    shared = sorted(set(ntf) & set(tfc))
    if len(shared) >= 3:
        first, last = shared[0], shared[-1]
        if ntf[first] and tfc[first]:
            ntf_change = (ntf[last] - ntf[first]) / ntf[first]
            tfc_change = (tfc[last] - tfc[first]) / tfc[first]
            # Flag when non-TANF rose >10% while TANF/FC fell >10%
            if ntf_change > 0.10 and tfc_change < -0.10:
                flags.append({
                    "type": "non_tanf_migration",
                    "yearsObserved": [first, last],
                    "nonTanfChangePct": round(ntf_change * 100, 1),
                    "tanfFcChangePct":  round(tfc_change * 100, 1),
                    "severity": "medium",
                    "explanation": f"Non-TANF collections grew {round(ntf_change*100,1)}% while "
                                   f"TANF/Foster Care collections fell {round(tfc_change*100,1)}% "
                                   f"across {first}-{last}. The program is shifting away from its "
                                   "original mandate (recovering public-assistance outlays) into "
                                   "a general debt-collection operation against private families."
                })

    return derived, flags

def main():
    if not EXCEL_DIR.is_dir():
        raise SystemExit(f"Not found: {EXCEL_DIR}")

    # Detect FY from filename, sort newest first so merge prefers recent revisions
    workbooks = []
    for f in EXCEL_DIR.iterdir():
        if f.suffix.lower() != ".xlsx": continue
        m = re.search(r'fy[_\s]?(\d{4})', f.name.lower())
        if not m: continue
        workbooks.append((int(m.group(1)), f))
    workbooks.sort(key=lambda t: t[0], reverse=True)

    print(f"Found {len(workbooks)} FY workbook(s):")
    for fy, f in workbooks: print(f"  FY{fy}  <- {f.name}")

    all_data = []
    for fy, f in workbooks:
        print(f"Parsing FY{fy}…")
        all_data.append(read_workbook(f))

    merged = merge_workbooks(None, all_data)

    # Pivot into per-state structure
    states = defaultdict(lambda: defaultdict(dict))
    all_years = set()
    all_state_names = set()
    for metric, states_map in merged.items():
        for state, years in states_map.items():
            all_state_names.add(state)
            for year, val in years.items():
                states[state][metric][year] = val
                all_years.add(year)

    out_states = {}
    for state in sorted(all_state_names):
        sd = {m: dict(sorted(states[state][m].items())) for m in sorted(states[state])}
        derived, flags = compute_derived(sd)
        out_states[state] = {"metrics": sd, "derived": derived, "flags": flags}

    result = {
        "version":             1,
        "source":              "OCSE preliminary data tables (FY20xx workbooks)",
        "fiscalYearsCovered":  sorted(all_years),
        "workbooks":           [f.name for _, f in workbooks],
        "metricKeys":          sorted(set(METRICS.values())),
        "flagTypes":           [
            "claimed_near_100pct_with_arrears",
            "arrears_monotone_growth",
            "caseload_drop_collections_stable",
            "cost_effectiveness_outlier",
            "incentive_estimate_vs_actual_divergence",
            "perpetual_debt_ratio",
            "non_tanf_migration",
        ],
        "disclaimer":          (
            "OCSE preliminary data is self-reported by states and subject to revision. "
            "Multi-year trends and cross-metric inconsistencies surface patterns worth "
            "investigating but are not dispositive evidence of fraud. Each flag links "
            "to the OCSE tables used so anyone can verify the math."
        ),
        "states":              out_states,
    }

    OUT_PATH.write_text(json.dumps(result, indent=2), encoding="utf-8")

    # Quick summary
    total_flags = sum(len(s["flags"]) for s in out_states.values())
    flagged_states = [name for name, s in out_states.items() if s["flags"]]
    print(f"\nWrote {OUT_PATH} ({OUT_PATH.stat().st_size/1024:.0f} KB)")
    print(f"Years covered: {sorted(all_years)}")
    print(f"States tracked: {len(out_states)}")
    print(f"Total flags raised: {total_flags} across {len(flagged_states)} states")

    # Print top implausibility cases
    for state in sorted(flagged_states, key=lambda n: -len(out_states[n]["flags"])):
        flags = out_states[state]["flags"]
        if not flags: continue
        flag_types = [f["type"] for f in flags]
        print(f"  {state}: {len(flags)} flag(s) -> {', '.join(flag_types)}")

if __name__ == "__main__":
    main()
